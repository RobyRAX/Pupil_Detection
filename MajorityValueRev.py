import openpyxl
from openpyxl.chart import BarChart, Reference

def find_majority_value(column_values):
    if not column_values:
        return None

    majority_value = max(set(column_values), key=column_values.count)
    count_majority = column_values.count(majority_value)
    if count_majority >= 2:
        return majority_value, count_majority
    else:
        return 'All Differ'
# Load the existing workbook
workbook = openpyxl.load_workbook('Hasil/Hasil.xlsx')

# Get the desired sheet
sheet = workbook['Normal']

# Define the range of rows where your data is located
start_row = 3
end_row = 602

col = 16

sheet.cell(row=start_row - 2, column=col, value='Analysis')
sheet.cell(row=start_row - 1, column=col, value='Majority Pick')
# sheet.cell(row=start_row - 1, column=col, value='Majority Value Count')

# Iterate over each row
for row in range(start_row, end_row + 1):
    # Get the values from the three cells in the row (B3, E3, H3)
    values = [sheet.cell(row=row, column=2).value, 
              sheet.cell(row=row, column=5).value, 
              sheet.cell(row=row, column=8).value,
              sheet.cell(row=row, column=11).value,
              sheet.cell(row=row, column=14).value]
    
    # Calculate the majority value using the function with a default value of None
    majority_value = find_majority_value(values)
    
    # Write the majority value to the next column (I column in this case)
    sheet.cell(row=row, column=col).value = majority_value[0]
    sheet.cell(row=row, column=col + 1).value = majority_value[1]

# Create a bar chart
chart = BarChart()
# Set the data range for the chart
data = Reference(sheet, min_col=col, min_row=2, max_col=col, max_row=end_row)
# Add data and categories to the chart
chart.add_data(data, titles_from_data=True)
# Add the chart to the worksheet
sheet.add_chart(chart, "R81")
chart.title = "MajorityValue"

# Save the workbook
workbook.save('Hasil/Hasil.xlsx')

print('Majority Data Complete')