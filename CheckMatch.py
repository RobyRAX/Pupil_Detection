import openpyxl

def compare_columns(sheet, sheet_2, row, column1, column2):
    # Get the values from the two cells in the row
    value1 = sheet.cell(row=row, column=column1).value
    value2 = sheet_2.cell(row=row, column=column2).value

    # Compare the values
    if value1 == value2:
        return 'Same'
    else:
        return 'No'

# Load the existing workbook
workbook = openpyxl.load_workbook('Hasil/Hasil.xlsx')

# Get the desired sheet
sheet = workbook['Sheet']
sheet_2 = workbook['Autis']

# Define the range of rows where your data is located
start_row = 3
end_row = 602

# Define the columns to compare
column1 = 16  # Column B
column2 = 2  # Column C

sheet_2.cell(row=start_row - 2, column=column2 + 2, value='Analysis')
sheet_2.cell(row=start_row - 1, column=column2 + 2, value='Match Analysis')
sheet_2.cell(row=start_row - 1, column=column2 + 3, value='Match Result')

# Count the number of 'Same' values
same_count = 0

# Iterate over each row
for row in range(start_row, end_row + 1):
    # Compare the cell values in the row
    result = compare_columns(sheet,sheet_2, row, column1, column2)
    
    # Write the result to the next column (D column in this case)
    sheet_2.cell(row=row, column=column2 + 2).value = result

    # Check if the result is 'Same' and increment the count
    if result == 'Same':
        same_count += 1

# Calculate the percentage of 'Same' values
total_rows = end_row - start_row + 1
percentage_same = (same_count / total_rows) * 100

# Write the percentage to the cell (you can choose any cell for this)
count_cell = sheet_2.cell(row=start_row, column=column2 + 3)
count_cell.value = f'Same Count = {same_count}'
percentage_cell = sheet_2.cell(row=start_row + 1, column=column2 + 3)
percentage_cell.value = f'Same Percentage = {percentage_same:.2f}%'

# Save the workbook
workbook.save('Hasil/Hasil.xlsx')

print('Match Analysis Complete')
