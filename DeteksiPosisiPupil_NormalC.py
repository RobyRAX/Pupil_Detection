import os
import sys
import math
import cv2
import mediapipe as mp
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference

mp_drawing = mp.solutions.drawing_utils
mp_drawing_styles = mp.solutions.drawing_styles
mp_face_mesh = mp.solutions.face_mesh

def calculate_intersection(line1, line2):
    x1, y1 = line1[0]
    x2, y2 = line1[1]
    x3, y3 = line2[0]
    x4, y4 = line2[1]

    denominator = ((x1 - x2) * (y3 - y4)) - ((y1 - y2) * (x3 - x4))

    if denominator != 0:
        px = ((x1 * y2 - y1 * x2) * (x3 - x4) - (x1 - x2) * (x3 * y4 - y3 * x4)) / denominator
        py = ((x1 * y2 - y1 * x2) * (y3 - y4) - (y1 - y2) * (x3 * y4 - y3 * x4)) / denominator
        intersection_point = (int(px), int(py))
        # cv2.circle(image, intersection_point, 3, (0, 0, 255), -1)
        return intersection_point
    else:
        return None

def is_point_inside_quadrilateral(point, quad_points):
    x, y = point

    x_coords = [p[0] for p in quad_points]
    y_coords = [p[1] for p in quad_points]

    # Calculate the four edge vectors of the quadrilateral
    edge_vectors = [
        (quad_points[i][0] - quad_points[(i + 1) % 4][0], quad_points[i][1] - quad_points[(i + 1) % 4][1])
        for i in range(4)
    ]

    # Calculate the four vectors from the point to each quadrilateral vertex
    point_vectors = [
        (x - quad_points[i][0], y - quad_points[i][1])
        for i in range(4)
    ]

    # Check if the point is on the same side of each edge as the corresponding vertex
    for i in range(4):
        cross_product = edge_vectors[i][0] * point_vectors[i][1] - edge_vectors[i][1] * point_vectors[i][0]
        if cross_product > 0:
            return False

    return True

def read_images_from_folder(folder_path):
    images = []
    for filename in os.listdir(folder_path):
        img_path = os.path.join(folder_path, filename)
        if os.path.isfile(img_path):
            image = cv2.imread(img_path)
            if image is not None:
                images.append(image)
    return images

def read_images_filenames_from_folder(folder_path):
    images = []
    filenames = []
    for filename in os.listdir(folder_path):
        img_path = os.path.join(folder_path, filename)
        if os.path.isfile(img_path):
            img = cv2.imread(img_path)
            if img is not None:
                images.append(img)
                filenames.append(filename)
    return images, filenames

def clear_saved_text(file_path):
  orig_stdout = sys.stdout
  with open(file_path, 'w') as f:
      sys.stdout = f
      # Print your terminal messages here
      print("Hasil Pengujian")
  sys.stdout = orig_stdout

def save_terminal_output(file_path, message):
  orig_stdout = sys.stdout
  with open(file_path, 'a') as f:
      sys.stdout = f
      # Print your terminal messages here
      print(message)
  sys.stdout = orig_stdout

pupil_position_grid_index = 0

folder_path = 'Normal - C'
result_path = 'Hasil/Hasil_Normal_C.txt'

row = 1
col = 7

# Load the existing workbook
workbook = openpyxl.load_workbook('Hasil/Hasil.xlsx')
# Get the active sheet
sheet = workbook['Normal']

# Add the header row to the sheet
sheet.cell(row=row, column=col, value=folder_path)
# Define the header row
header_row = ['File Name', 'Pupil Position']
# Add the header row to the sheet
sheet.cell(row=row + 1, column=col, value=header_row[0])
sheet.cell(row=row + 1, column=col + 1, value=header_row[1])

clear_saved_text(result_path)

row_count = 0

with mp_face_mesh.FaceMesh(
  max_num_faces=1,
  refine_landmarks=True,
  min_detection_confidence=0.5,
  min_tracking_confidence=0.5) as face_mesh:

  for filename in os.listdir(folder_path):
    file_path = os.path.join(folder_path, filename)

    image = cv2.imread(file_path) 

    # To improve performance, optionally mark the image as not writeable to
    # pass by reference.
    image.flags.writeable = False
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(image)

    # Draw the face mesh annotations on the image.
    image.flags.writeable = True
    image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

    if results.multi_face_landmarks:
      for face_landmarks in results.multi_face_landmarks:
        right_eye_out = face_landmarks.landmark[33]           #Kelopak Luar
        right_eye_out_x = int(right_eye_out.x * image.shape[1])
        right_eye_out_y = int(right_eye_out.y * image.shape[0])

        right_eye_in = face_landmarks.landmark[133]           #Kelopak Dalam
        right_eye_in_x = int(right_eye_in.x * image.shape[1])
        right_eye_in_y = int(right_eye_in.y * image.shape[0])

        horizon_line = ((right_eye_out_x, right_eye_out_y), (right_eye_in_x, right_eye_in_y))
        
        right_eye_pupil = face_landmarks.landmark[468]           #Pupil Tengah
        right_eye_pupil_x = int(right_eye_pupil.x * image.shape[1])
        right_eye_pupil_y = int(right_eye_pupil.y * image.shape[0])

        right_pupil = (right_eye_pupil_x, right_eye_pupil_y)

        # Calculate the length and angle of the original line
        length = math.sqrt((right_eye_in_x - right_eye_out_x) ** 2 + (right_eye_in_y - right_eye_out_y) ** 2)
        angle = math.atan2(right_eye_in_y - right_eye_out_y, right_eye_in_x - right_eye_out_x)

        midpoint_right_eye = (
          (right_eye_out_x + right_eye_in_x) // 2, 
          (right_eye_out_y + right_eye_in_y) // 2
        )

        midpoint_right_eye_out = (
          (right_eye_out_x + midpoint_right_eye[0]) // 2, 
          (right_eye_out_y + midpoint_right_eye[1]) // 2
        )

        midpoint_right_eye_in = (
          (right_eye_in_x + midpoint_right_eye[0]) // 2, 
          (right_eye_in_y + midpoint_right_eye[1]) // 2
        )

        # Set the length of the perpendicular line
        perpendicular_length = math.sqrt((face_landmarks.landmark[470].x * image.shape[0] - 
                                          face_landmarks.landmark[472].x * image.shape[0])**2 + 
                                          (face_landmarks.landmark[470].y * image.shape[1] - 
                                          face_landmarks.landmark[472].y * image.shape[1])**2)

        # Calculate the endpoint of the perpendicular line
        perpendicular_endpoint_1_b = (
            int(midpoint_right_eye[0] + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(midpoint_right_eye[1] + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_1_t = (
            int(midpoint_right_eye[0] + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(midpoint_right_eye[1] + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_mid = (perpendicular_endpoint_1_b, perpendicular_endpoint_1_t)

        perpendicular_endpoint_2_b = (
            int(midpoint_right_eye_out[0] + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(midpoint_right_eye_out[1] + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_2_t = (
            int(midpoint_right_eye_out[0] + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(midpoint_right_eye_out[1] + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_midout = (perpendicular_endpoint_2_b, perpendicular_endpoint_2_t)

        perpendicular_endpoint_3_b = (
            int(midpoint_right_eye_in[0] + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(midpoint_right_eye_in[1] + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_3_t = (
            int(midpoint_right_eye_in[0] + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(midpoint_right_eye_in[1] + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_midin = (perpendicular_endpoint_3_b, perpendicular_endpoint_3_t)

        perpendicular_endpoint_4_b = (
            int(right_eye_in_x + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(right_eye_in_y + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_4_t = (
            int(right_eye_in_x + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(right_eye_in_y + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_in = (perpendicular_endpoint_4_b, perpendicular_endpoint_4_t)

        perpendicular_endpoint_5_b = (
            int(right_eye_out_x + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(right_eye_out_y + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_5_t = (
            int(right_eye_out_x + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(right_eye_out_y + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_out = (perpendicular_endpoint_5_b, perpendicular_endpoint_5_t)

        #--------------------
        # Set the offset of the horizontal line
        horizon_offset = math.sqrt((midpoint_right_eye[0] - midpoint_right_eye_out[0]) ** 2 + (midpoint_right_eye[1] - midpoint_right_eye_out[1]) ** 2)
        # Calculate the offset values
        offset_x = horizon_offset * math.sin(angle)
        offset_y = horizon_offset * math.cos(angle)

        # Calculate the start and end points of the offset line
        offset_b_line_start = (int(right_eye_out_x - offset_x), int(right_eye_out_y + offset_y))
        offset_b_line_end = (int(right_eye_in_x - offset_x), int(right_eye_in_y + offset_y))
        horizon_bot_line = ((offset_b_line_start), (offset_b_line_end))

        offset_t_line_start = (int(right_eye_out_x + offset_x), int(right_eye_out_y - offset_y))
        offset_t_line_end = (int(right_eye_in_x + offset_x), int(right_eye_in_y - offset_y))
        horizon_top_line = [(offset_t_line_start), (offset_t_line_end)]

        pt1 = calculate_intersection(horizon_line, perpendicular_mid)
        pt2 = calculate_intersection(horizon_line, perpendicular_midout)
        pt3 = calculate_intersection(horizon_line, perpendicular_midin)

        pt4 = calculate_intersection(horizon_top_line, perpendicular_midout)
        pt5 = calculate_intersection(horizon_top_line, perpendicular_mid)
        pt6 = calculate_intersection(horizon_top_line, perpendicular_midin)

        pt7 = calculate_intersection(horizon_bot_line, perpendicular_mid)
        pt8 = calculate_intersection(horizon_bot_line, perpendicular_midout)
        pt9 = calculate_intersection(horizon_bot_line, perpendicular_midin)

        pt10 = calculate_intersection(horizon_top_line, perpendicular_out)
        pt11 = (right_eye_out_x, right_eye_out_y)
        pt12 = calculate_intersection(horizon_bot_line, perpendicular_out)

        pt13 = calculate_intersection(horizon_top_line, perpendicular_in)
        pt14 = (right_eye_in_x, right_eye_in_y)
        pt15 = calculate_intersection(horizon_bot_line, perpendicular_in)

        quad_1 = (pt1, pt2, pt4, pt5)
        quad_2 = (pt3, pt1, pt5, pt6)
        quad_3 = (pt9, pt7, pt1, pt3)
        quad_4 = (pt7, pt8, pt2, pt1)
        quad_5 = (pt11, pt2, pt4, pt10)
        quad_6 = (pt12, pt8, pt2, pt11)
        quad_7 = (pt9, pt15, pt14, pt3)
        quad_8 = (pt3, pt14, pt13, pt6)

        if is_point_inside_quadrilateral(right_pupil, quad_1):
          pupil_position_grid_index = 1
        elif is_point_inside_quadrilateral(right_pupil, quad_2):
          pupil_position_grid_index = 2
        elif is_point_inside_quadrilateral(right_pupil, quad_3):
          pupil_position_grid_index = 6
        elif is_point_inside_quadrilateral(right_pupil, quad_4):
          pupil_position_grid_index = 5
        elif is_point_inside_quadrilateral(right_pupil, quad_5):
          pupil_position_grid_index = 0
        elif is_point_inside_quadrilateral(right_pupil, quad_6):
          pupil_position_grid_index = 4
        elif is_point_inside_quadrilateral(right_pupil, quad_7):
          pupil_position_grid_index = 7
        elif is_point_inside_quadrilateral(right_pupil, quad_8):
          pupil_position_grid_index = 3

      result_message = 'FileName = {} // PupilPosition = {}'.format(filename, pupil_position_grid_index)
      print(result_message)
      save_terminal_output(result_path, result_message)

      # Data to fill in the rows
      data = [filename, pupil_position_grid_index]
      sheet.cell(row=row + 2 + row_count, column=col, value=data[0])
      sheet.cell(row=row + 2 + row_count, column=col + 1, value=data[1])
      row_count += 1
    else:
       # Data to fill in the rows
      data = [filename, 'Null']
      sheet.cell(row=row + 2 + row_count, column=col, value=data[0])
      sheet.cell(row=row + 2 + row_count, column=col + 1, value=data[1])
      row_count += 1
# Create a bar chart
chart = BarChart()
# Set the data range for the chart
data = Reference(sheet, min_col=col + 1, min_row=2, max_col=col + 1, max_row=row_count + 2)
# Set the categories for the chart
categories = Reference(sheet, min_col=col, min_row=3, max_row=row_count + 2)
# Add data and categories to the chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
# Add the chart to the worksheet
sheet.add_chart(chart, "R33")
chart.title = folder_path

# Save the workbook
workbook.save('Hasil/Hasil.xlsx')

#---------------------------------------------------------------------------------------------------------------

images, filenames = read_images_filenames_from_folder(folder_path)
current_image_index = 0
while True:
  with mp_face_mesh.FaceMesh(
    max_num_faces=1,
    refine_landmarks=True,
    min_detection_confidence=0.5,
    min_tracking_confidence=0.5) as face_mesh:

    image = images[current_image_index]
    filename = filenames[current_image_index]
    cv2.putText(image, filename, (0, image.shape[0]-20), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

    # To improve performance, optionally mark the image as not writeable to
    # pass by reference.
    image.flags.writeable = False
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(image)

    # Draw the face mesh annotations on the image.
    image.flags.writeable = True
    image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

    # Define grid properties
    grid_rows = 2
    grid_cols = 4
    grid_color = (0, 255, 0)  # Green color (BGR format)
    grid_thickness = 2  # Line thickness

    iteration = 0

    # Calculate cell width and height
    cell_width = image.shape[1] // 15
    cell_height = cell_width

    # Draw the grid
    for row in range(grid_rows):
      for col in range(grid_cols):
        # Calculate the top-left and bottom-right corners of the cell
        x1 = col * cell_width
        y1 = row * cell_height
        x2 = (col + 1) * cell_width
        y2 = (row + 1) * cell_height

        # Draw a rectangle for each cell
        cv2.rectangle(image, (x1, y1), (x2, y2), grid_color, grid_thickness)
        cv2.putText(image, str(iteration), (x1 + cell_height//4, y2 - cell_height//4), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 0), 2)
        iteration += 1

    if results.multi_face_landmarks:
      for face_landmarks in results.multi_face_landmarks:
        right_eye_out = face_landmarks.landmark[33]           #Kelopak Luar
        right_eye_out_x = int(right_eye_out.x * image.shape[1])
        right_eye_out_y = int(right_eye_out.y * image.shape[0])
        cv2.circle(image, (right_eye_out_x, right_eye_out_y), 3, (255, 0, 0), -1)

        right_eye_in = face_landmarks.landmark[133]           #Kelopak Dalam
        right_eye_in_x = int(right_eye_in.x * image.shape[1])
        right_eye_in_y = int(right_eye_in.y * image.shape[0])
        cv2.circle(image, (right_eye_in_x, right_eye_in_y), 3, (255, 0, 0), -1)

        horizon_line = ((right_eye_out_x, right_eye_out_y), (right_eye_in_x, right_eye_in_y))
        
        right_eye_pupil = face_landmarks.landmark[468]           #Pupil Tengah
        right_eye_pupil_x = int(right_eye_pupil.x * image.shape[1])
        right_eye_pupil_y = int(right_eye_pupil.y * image.shape[0])

        right_pupil = (right_eye_pupil_x, right_eye_pupil_y)

        # Calculate the length and angle of the original line
        length = math.sqrt((right_eye_in_x - right_eye_out_x) ** 2 + (right_eye_in_y - right_eye_out_y) ** 2)
        angle = math.atan2(right_eye_in_y - right_eye_out_y, right_eye_in_x - right_eye_out_x)

        midpoint_right_eye = (
          (right_eye_out_x + right_eye_in_x) // 2, 
          (right_eye_out_y + right_eye_in_y) // 2
        )

        midpoint_right_eye_out = (
          (right_eye_out_x + midpoint_right_eye[0]) // 2, 
          (right_eye_out_y + midpoint_right_eye[1]) // 2
        )

        midpoint_right_eye_in = (
          (right_eye_in_x + midpoint_right_eye[0]) // 2, 
          (right_eye_in_y + midpoint_right_eye[1]) // 2
        )

        # Set the length of the perpendicular line
        perpendicular_length = math.sqrt((face_landmarks.landmark[470].x * image.shape[0] - 
                                          face_landmarks.landmark[472].x * image.shape[0])**2 + 
                                          (face_landmarks.landmark[470].y * image.shape[1] - 
                                          face_landmarks.landmark[472].y * image.shape[1])**2)

        # Calculate the endpoint of the perpendicular line
        perpendicular_endpoint_1_b = (
            int(midpoint_right_eye[0] + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(midpoint_right_eye[1] + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_1_t = (
            int(midpoint_right_eye[0] + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(midpoint_right_eye[1] + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_mid = (perpendicular_endpoint_1_b, perpendicular_endpoint_1_t)

        perpendicular_endpoint_2_b = (
            int(midpoint_right_eye_out[0] + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(midpoint_right_eye_out[1] + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_2_t = (
            int(midpoint_right_eye_out[0] + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(midpoint_right_eye_out[1] + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_midout = (perpendicular_endpoint_2_b, perpendicular_endpoint_2_t)

        perpendicular_endpoint_3_b = (
            int(midpoint_right_eye_in[0] + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(midpoint_right_eye_in[1] + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_3_t = (
            int(midpoint_right_eye_in[0] + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(midpoint_right_eye_in[1] + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_midin = (perpendicular_endpoint_3_b, perpendicular_endpoint_3_t)

        perpendicular_endpoint_4_b = (
            int(right_eye_in_x + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(right_eye_in_y + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_4_t = (
            int(right_eye_in_x + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(right_eye_in_y + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_in = (perpendicular_endpoint_4_b, perpendicular_endpoint_4_t)

        perpendicular_endpoint_5_b = (
            int(right_eye_out_x + perpendicular_length * math.cos(angle + math.pi / 2)),
            int(right_eye_out_y + perpendicular_length * math.sin(angle + math.pi / 2))
        )
        perpendicular_endpoint_5_t = (
            int(right_eye_out_x + perpendicular_length * math.cos(angle - math.pi / 2)),
            int(right_eye_out_y + perpendicular_length * math.sin(angle - math.pi / 2))
        )
        perpendicular_out = (perpendicular_endpoint_5_b, perpendicular_endpoint_5_t)

        #--------------------
        # Set the offset of the horizontal line
        horizon_offset = math.sqrt((midpoint_right_eye[0] - midpoint_right_eye_out[0]) ** 2 + (midpoint_right_eye[1] - midpoint_right_eye_out[1]) ** 2)
        # Calculate the offset values
        offset_x = horizon_offset * math.sin(angle)
        offset_y = horizon_offset * math.cos(angle)

        # Calculate the start and end points of the offset line
        offset_b_line_start = (int(right_eye_out_x - offset_x), int(right_eye_out_y + offset_y))
        offset_b_line_end = (int(right_eye_in_x - offset_x), int(right_eye_in_y + offset_y))
        horizon_bot_line = ((offset_b_line_start), (offset_b_line_end))

        offset_t_line_start = (int(right_eye_out_x + offset_x), int(right_eye_out_y - offset_y))
        offset_t_line_end = (int(right_eye_in_x + offset_x), int(right_eye_in_y - offset_y))
        horizon_top_line = [(offset_t_line_start), (offset_t_line_end)]

        pt1 = calculate_intersection(horizon_line, perpendicular_mid)
        pt2 = calculate_intersection(horizon_line, perpendicular_midout)
        pt3 = calculate_intersection(horizon_line, perpendicular_midin)

        pt4 = calculate_intersection(horizon_top_line, perpendicular_midout)
        pt5 = calculate_intersection(horizon_top_line, perpendicular_mid)
        pt6 = calculate_intersection(horizon_top_line, perpendicular_midin)

        pt7 = calculate_intersection(horizon_bot_line, perpendicular_mid)
        pt8 = calculate_intersection(horizon_bot_line, perpendicular_midout)
        pt9 = calculate_intersection(horizon_bot_line, perpendicular_midin)

        pt10 = calculate_intersection(horizon_top_line, perpendicular_out)
        pt11 = (right_eye_out_x, right_eye_out_y)
        pt12 = calculate_intersection(horizon_bot_line, perpendicular_out)

        pt13 = calculate_intersection(horizon_top_line, perpendicular_in)
        pt14 = (right_eye_in_x, right_eye_in_y)
        pt15 = calculate_intersection(horizon_bot_line, perpendicular_in)

        quad_1 = (pt1, pt2, pt4, pt5)
        quad_2 = (pt3, pt1, pt5, pt6)
        quad_3 = (pt9, pt7, pt1, pt3)
        quad_4 = (pt7, pt8, pt2, pt1)
        quad_5 = (pt11, pt2, pt4, pt10)
        quad_6 = (pt12, pt8, pt2, pt11)
        quad_7 = (pt9, pt15, pt14, pt3)
        quad_8 = (pt3, pt14, pt13, pt6)

        poly_thickness = 2
        cv2.polylines(image, [np.array(quad_1)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_2)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_3)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_4)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_5)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_6)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_7)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)
        cv2.polylines(image, [np.array(quad_8)], isClosed=True, color=(0, 255, 0), thickness=poly_thickness)

        cv2.circle(image, right_pupil, 5, (255, 255, 255), -1)

        if is_point_inside_quadrilateral(right_pupil, quad_1):
          cv2.rectangle(image, (1 * cell_width, 0 * cell_height), (2 * cell_width, 1 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 1
        elif is_point_inside_quadrilateral(right_pupil, quad_2):
          cv2.rectangle(image, (2 * cell_width, 0 * cell_height), (3 * cell_width, 1 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 2
        elif is_point_inside_quadrilateral(right_pupil, quad_3):
          cv2.rectangle(image, (2 * cell_width, 1 * cell_height), (3 * cell_width, 2 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 6
        elif is_point_inside_quadrilateral(right_pupil, quad_4):
          cv2.rectangle(image, (1 * cell_width, 1 * cell_height), (2 * cell_width, 2 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 5
        elif is_point_inside_quadrilateral(right_pupil, quad_5):
          cv2.rectangle(image, (0 * cell_width, 0 * cell_height), (1 * cell_width, 1 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 0
        elif is_point_inside_quadrilateral(right_pupil, quad_6):
          cv2.rectangle(image, (0 * cell_width, 1 * cell_height), (1 * cell_width, 2 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 4
        elif is_point_inside_quadrilateral(right_pupil, quad_7):
          cv2.rectangle(image, (3 * cell_width, 1 * cell_height), (4 * cell_width, 2 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 7
        elif is_point_inside_quadrilateral(right_pupil, quad_8):
          cv2.rectangle(image, (3 * cell_width, 0 * cell_height), (4 * cell_width, 1 * cell_height), grid_color, cv2.FILLED)
          pupil_position_grid_index = 3
    
    # cv2.namedWindow('Deteksi Posisi Pupil', cv2.WINDOW_NORMAL)
    # cv2.resizeWindow('Deteksi Posisi Pupil', 800, 450)
    cv2.imshow('Deteksi Posisi Pupil', image)

    # Wait for key press
    key = cv2.waitKey(0)

    # Exit if 'Esc' key is pressed
    if key == 27 or key == ord('q'):
        break

    # Move to the next image if 'Right arrow' key is pressed
    if key == 83 or key == 100:
        current_image_index = (current_image_index + 1) % len(images)

    # Move to the previous image if 'Left arrow' key is pressed
    if key == 81 or key == 97:
        current_image_index = (current_image_index - 1) % len(images)

cv2.destroyAllWindows()